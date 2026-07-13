# -*- coding: utf-8 -*-
"""
価格改定で使うマスタの読込。

1) 送料・資材マスタ（項目1のサイズコード → 送料/資材/配送種別）
   一次ソース: 価格変更スプレッドシート（Secrets PRICING_SHEET_ID、リンク共有のxlsx export）。
   取得失敗・未設定時は同梱の lib/pricing/data/cost_master.csv にフォールバック。
   シート側（「送料表」D:E・「費用」E:F）を直せばアプリも追従する。

2) NE商品マスタ（JAN→商品コード・売価・原価・項目1）
   共有 master.csv（汎用マスタ変換）には売価・原価が無いため、価格改定用に
   NEカスタムCSV（売価・原価入り）を Drive に pricing_ne_master.csv として保存・再利用する。
   Drive まわりは請求書モジュールの drive_master をそのまま使う。
"""
import io
import os
import unicodedata

import pandas as pd

from lib.invoice import csv_import, drive_master

# Drive上の価格改定用NE商品マスタの固定ファイル名（フォルダはPRODUCT_MASTER_FOLDER_ID）
PRICING_MASTER_NAME = "pricing_ne_master.csv"

_DATA_DIR = os.path.join(os.path.dirname(__file__), "data")

# 列名のゆらぎ → 正規名
_COL_ALIASES = {
    "JANコード": ("JANコード", "JAN", "jan", "JANCD", "JANcd"),
    "商品コード": ("商品コード", "商品CD", "商品cd"),
    "売価": ("売価", "販売価格", "商品価格", "NE売価"),
    "原価": ("原価", "仕入原価", "下代", "NE原価", "仕入価格"),
    "項目1": ("項目1", "項目１"),
}


def norm_key(value):
    """サイズコード/JAN/商品コードの正規化: NFKC・前後空白・末尾".0"除去・小文字はそのまま。"""
    s = unicodedata.normalize("NFKC", str(value)).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return s


def _norm_columns(df):
    """列名をNFKC正規化し、ゆらぎを正規名に寄せる。"""
    df = df.rename(columns={c: unicodedata.normalize("NFKC", str(c)).strip() for c in df.columns})
    ren = {}
    for canon, aliases in _COL_ALIASES.items():
        if canon in df.columns:
            continue
        for a in aliases:
            if a in df.columns:
                ren[a] = canon
                break
    return df.rename(columns=ren)


# ── 送料・資材マスタ ──────────────────────────────────────

def _mail_or_takuhai(key):
    """項目1コードから配送種別を推定（数値サイズ=宅配便、メール便系コード=メール便）。"""
    return "メール便" if key in ("nekop", "yuup1", "yuup2", "yuup3", "1", "3") else "宅配便"


def load_cost_master_from_sheet(sheet_id):
    """スプレッドシートのxlsx exportから「送料表」D:E・「費用」E:F を読み、DataFrameを返す。"""
    import requests
    from openpyxl import load_workbook
    url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
    res = requests.get(url, timeout=30)
    res.raise_for_status()
    wb = load_workbook(io.BytesIO(res.content), data_only=True, read_only=True)

    def pairs(sheet, key_col, val_col):
        ws = wb[sheet]
        d = {}
        for row in ws.iter_rows(min_col=key_col, max_col=val_col):
            k, v = row[0].value, row[-1].value
            if k is None or v is None or not isinstance(v, (int, float)):
                continue
            key = norm_key(k)
            if key and key not in d:  # VLOOKUPと同じく最初の一致を採用
                d[key] = float(v)
        return d

    shipping = pairs("送料表", 4, 5)   # D:E サイズ→平均運賃
    material = pairs("費用", 5, 6)     # E:F サイズ→資材費
    shipping.pop("値", None)
    keys = list(shipping.keys()) + [k for k in material if k not in shipping]
    rows = [{"項目1": k, "送料": shipping.get(k), "資材": material.get(k),
             "配送種別": _mail_or_takuhai(k)} for k in keys]
    df = pd.DataFrame(rows)
    if df.empty or df["送料"].isna().all():
        raise ValueError("シートから送料マスタを取得できませんでした")
    return df


def load_cost_master_bundled():
    """同梱CSV（シートから抽出したスナップショット）を読む。"""
    path = os.path.join(_DATA_DIR, "cost_master.csv")
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    df["項目1"] = df["項目1"].map(norm_key)
    for c in ("送料", "資材"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def cost_lookup(df):
    """DataFrame → {項目1: (送料, 資材, 配送種別)}。送料/資材はNaN→None。"""
    table = {}
    for _, r in df.iterrows():
        ship = r["送料"]
        mat = r["資材"]
        table[norm_key(r["項目1"])] = (
            None if pd.isna(ship) else float(ship),
            None if pd.isna(mat) else float(mat),
            str(r.get("配送種別", "")) or _mail_or_takuhai(norm_key(r["項目1"])),
        )
    return table


# ── NE商品マスタ ─────────────────────────────────────────

def load_ne_master(file_bytes):
    """NEカスタム(商品マスタ)CSVを読む。商品コード必須。売価・原価・JAN・項目1は有無を検査して返す。"""
    df = _norm_columns(csv_import.read_csv_auto(file_bytes))
    if "商品コード" not in df.columns:
        raise ValueError(f"必須列「商品コード」が見つかりません / 実際の列: {list(df.columns)}")
    missing = [c for c in ("JANコード", "売価", "原価", "項目1") if c not in df.columns]
    return df, missing


def find_drive_master(folder_id):
    """Drive上の pricing_ne_master.csv を探す（無ければNone）。"""
    return drive_master.find_file(PRICING_MASTER_NAME, folder_id)


def download_drive_master(file_id):
    return drive_master.download_bytes(file_id)


def save_drive_master(file_bytes, folder_id):
    """pricing_ne_master.csv を上書き保存（無ければ新規作成）。"""
    return drive_master.upload_or_replace(file_bytes, PRICING_MASTER_NAME, folder_id)


def build_lookup(ne_df):
    """
    NE商品マスタ → 突合用のインデックスを作る。
    返り値: (jan→商品コード dict, 商品コード→{売価,原価,項目1,商品名} dict)
    """
    cols = ne_df.columns
    jan_map = {}
    info = {}
    for _, r in ne_df.iterrows():
        code = norm_key(r["商品コード"])
        if not code or code == "nan":
            continue
        if "JANコード" in cols:
            jan = norm_key(r["JANコード"])
            if jan and jan != "nan" and jan not in jan_map:
                jan_map[jan] = code
        info[code.lower()] = {
            "商品コード": code,
            "商品名": str(r.get("商品名", "") or ""),
            "売価": r.get("売価", ""),
            "原価": r.get("原価", ""),
            "項目1": norm_key(r.get("項目1", "")) if "項目1" in cols else "",
        }
    return jan_map, info
